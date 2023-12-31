# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'BradScraper.ui'
#
# Created by: PyQt5 UI code generator 5.15.9
#
# WARNING: Any manual changes made to this file will be lost when pyuic5 is
# run again.  Do not edit this file unless you know what you are doing.

from PyQt5 import QtCore, QtGui, QtWidgets 
from PyQt5.QtCore import QThread, pyqtSignal
from PyQt5.QtWidgets import QMessageBox, QWidget

# Global Variables
TIME_LIMIT = 100
entries = ['https://metrovac.com/collections/product']
model = QtGui.QStandardItemModel()
header = ['No', 'Product Name', 'Product Price', 'Image Src']
excel_file = 'Metro Vac Template.xlsx'
color_array = ['Pink', 'Blue', 'Orange', 'Red', 'Green', 'Yellow', 'Purple']
exceptional_array = ['2 - 1.17HP', '3 - 1.7HP', 'ES-105', 'ES-105T', 'ES-109T']

def getDescriptionAndURL(data):
    base_values = []
    for datum in data:
        script_contents="\n".join([script.string for script in datum if script.string])
        json_data = json.loads(script_contents)
        if 'description' in json_data:
            offers = json_data['offers']
            for offer in offers:
                row = []
                row.append(offer['sku'])
                # row.append(json_data['name']+" "+offer['sku'])
                descrb = json_data['description']
                descrb = descrb.replace('\n', '')
                descrb = descrb.strip()
                row.append(descrb)
                row.append(offer['price'])
                row.append(offer['url'])
                base_values.append(row)
            return base_values
    return []

def getVideoURL(video_url_data):
    for video_url_datum in video_url_data:
        if 'youtube' in video_url_datum['src']:
            str_video_url = video_url_datum['src']
            return str_video_url
    return ''

def getVariantsImageURLS(image_url_data):
    ## image url handle
    mainData = []
    for image_url_datum in image_url_data:
        script_contents="\n".join([script.string for script in image_url_datum if script.string])
        try:
            start_index = script_contents.index('product:')+len('product:')
            end_index = script_contents.index('currentVariant:')
            script_contents = script_contents[start_index:end_index]
            script_contents = script_contents.strip()
            script_contents = script_contents[:len(script_contents)-1]
            json_data = json.loads(script_contents)
            images_urls = json_data['images']
            variants_data = json_data['variants']
            for variants_datum in variants_data:
                data_row = []
                data_row.append(variants_datum['name'])
                data_row.append(variants_datum['option1'])
                data_row.append(variants_datum['option2'])
                data_row.append(variants_datum['option3'])
                data_row.append(images_urls)
                mainData.append(data_row)
            return mainData
        except ValueError:
            return []
    return []

def getTableValues(table):
    th_idx=0
    table_data = []
    try:
        for row in table.find_all('tr'):
            row_data = []
            if th_idx==0:
                for cell in row.find_all('th'):
                    row_data.append(cell.text)
            for cell in row.find_all('td'):
                row_data.append(cell.text)
            th_idx += 1
            table_data.append(row_data)
        
        return table_data
    except:
        return table_data

def getProperValueById(id, table, num):
    try:
        for tbl_idx in range(1,len(table)):
            if id == 'Attachments':
                if id in table[tbl_idx][0] or 'Accessories' in table[tbl_idx][0]:
                    return table[tbl_idx][num]
            if id == 'Weight':
                if id in table[tbl_idx][0] or 'Dimension' in table[tbl_idx][0]:
                    return table[tbl_idx][num]
            else:
                if id.lower() in table[tbl_idx][0].lower():
                    return table[tbl_idx][num]
    except:
        return ''
    
def getProperIdxInTable(table, variants):
    try:
        for tbl_row in table:
            for var_idx in range(1, 4):
                length = len(tbl_row)
                if length <= 2:
                    return 1
                for tbl_element_idx in range(1,length):
                    inner_var = variants[var_idx]
                    if inner_var == None:
                        continue
                    if inner_var.lower().strip() in 'ES-105'.strip():
                        print(inner_var)
                        return 1
                    if inner_var.lower().strip() in 'ES-105T'.strip():
                        print(inner_var)
                        return 2
                    if inner_var.lower().strip() in 'ES-109T'.strip():
                        print(inner_var)
                        return 3
                    if inner_var.lower().strip() in 'Single Speed'.strip():
                        print(inner_var)
                        return 1
                    if inner_var.lower().strip() in 'Two Speed'.strip():
                        print(inner_var)
                        return 2
                    inner_var = inner_var.split(' ')[0]

                    if inner_var in tbl_row[tbl_element_idx]:
                        return tbl_element_idx
        return 1
    except:
        return 1

def getProperImageLink(id, links):
    try:
        if id >= len(links):
            return ''
        else:
            return links[id]
    except:
        return ''

def getSize(targets):
    try:
        for idx in range(1,4):
            if 'Ft' in targets[idx]:
                return targets[idx]
            if 'In' in targets[idx]:
                return targets[idx]
            idx += 1
    except:
        return ''
    
def getMotorPower(targets):
    try:
        for idx in range(1,4):
            if 'HP' in targets[idx]:
                return targets[idx]
            idx += 1
    except:
        return ''
    
def getColor(targets):
    try:
        for idx in range(1,4):
            if targets[idx] in color_array:
                return targets[idx]
            idx += 1
    except:
        return ''

def setCellValue(cell, value):
    cell.value = value

    # Enable text wrapping for the cell
    cell.alignment = Alignment(wrapText=True)
class External(QThread):
    """
    Runs a counter thread.
    """
    countChanged = pyqtSignal(int)
    countFinished = pyqtSignal()
    countErrored = pyqtSignal()
    pageChanged = pyqtSignal(int)
    
    def run(self):
        try:
            # Making a GET request
            context = ssl._create_unverified_context()
                
            index = model.index(0, 0)
            item = model.data(index, QtCore.Qt.DisplayRole)
            workbook = load_workbook(excel_file)
            standard_sheet = None

            for page_number in range(1, 5):
                count = 0
                if page_number == 1:
                    sheet = workbook.worksheets[0]
                    sheet.title = 'page1'
                    sheet.freeze_panes = 'A2'
                    standard_sheet = sheet
                    sheet.column_dimensions['C'].width = 45
                    sheet.column_dimensions['P'].width = 30
                    sheet.column_dimensions['Q'].width = 20
                    sheet.column_dimensions['R'].width = 20
                    sheet.column_dimensions['S'].width = 20
                    sheet.column_dimensions['T'].width = 20
                    sheet.column_dimensions['U'].width = 20
                    sheet.column_dimensions['V'].width = 20
                    sheet.column_dimensions['W'].width = 20
                    sheet.column_dimensions['X'].width = 20
                else:
                    # Create a new sheet in the original workbook
                    sheet = workbook.copy_worksheet(standard_sheet)
                    sheet.title = 'page'+str(page_number)
                    start_row = 2
                    end_row = sheet.max_row
                    sheet.delete_rows(start_row, end_row)
                try:
                    r = requests.get(item + "?page=" + str( page_number ) )
                except:
                    self.countErrored.emit() 

                # Parsing the HTML
                soup = BeautifulSoup(r.content, 'html.parser')
                links = soup.find_all('a', class_='stretched-link')
                length = len(links)
                
                # try:
                
                row = 2

                for idx in range( length ):
                    url='https://metrovac.com/'+links[idx].get('href')
                    req = Request(
                        url, 
                        headers={'User-Agent': 'Mozilla/5.0'}
                    )
                    
                    page = urllib.request.urlopen(req,context=context)
                    inner_soup = BeautifulSoup(page, 'html.parser')

                    data = inner_soup.find_all('script', type='application/ld+json')
                    video_url_data = inner_soup.find_all('iframe')
                    image_url_data = inner_soup.find_all('script', id='four-shopify-base-script')
                    table = inner_soup.find('table', {'id': 'tablefield-0'})

                    ## video url handle
                    str_video_url = getVideoURL(video_url_data)
                    mainDataAndimages_urls = getVariantsImageURLS(image_url_data)
                    base_values = getDescriptionAndURL(data)
                    table_values = getTableValues(table)
                    index_main = 0

                    for base_row_value in base_values:
                        setCellValue(sheet.cell(row, 1), base_row_value[0])
                        setCellValue(sheet.cell(row, 2), mainDataAndimages_urls[index_main][0])
                        setCellValue(sheet.cell(row, 3), base_row_value[1])
                        setCellValue(sheet.cell(row, 4), base_row_value[2])
                        setCellValue(sheet.cell(row, 5), getColor(mainDataAndimages_urls[index_main]))
                        
                        alternatives = getProperIdxInTable(table_values, mainDataAndimages_urls[index_main] )
                        setCellValue(sheet.cell(row, 6), getProperValueById('Weight', table_values, alternatives))
                        setCellValue(sheet.cell(row, 7), getProperValueById('Construction', table_values, alternatives))
                        setCellValue(sheet.cell(row, 8), getProperValueById('Motor', table_values, alternatives))
                        setCellValue(sheet.cell(row, 9), getProperValueById('Amps', table_values, alternatives))
                        setCellValue(sheet.cell(row, 10), getProperValueById('Watts', table_values, alternatives))
                        setCellValue(sheet.cell(row, 11), getProperValueById('Air Volume', table_values, alternatives))
                        setCellValue(sheet.cell(row, 12), getProperValueById('Air Speed', table_values, alternatives))
                        setCellValue(sheet.cell(row, 13), getProperValueById('Airflow', table_values, alternatives))
                        setCellValue(sheet.cell(row, 14), getProperValueById('Cord', table_values, alternatives))
                        setCellValue(sheet.cell(row, 15), getProperValueById('Hose', table_values, alternatives))
                        setCellValue(sheet.cell(row, 16), getProperValueById('Attachments', table_values, alternatives))
                        setCellValue(sheet.cell(row, 17), getProperImageLink(0, mainDataAndimages_urls[index_main][4]))
                        setCellValue(sheet.cell(row, 18), getProperImageLink(1, mainDataAndimages_urls[index_main][4]))
                        setCellValue(sheet.cell(row, 19), getProperImageLink(2, mainDataAndimages_urls[index_main][4]))
                        setCellValue(sheet.cell(row, 20), getProperImageLink(3, mainDataAndimages_urls[index_main][4]))
                        setCellValue(sheet.cell(row, 21), getProperImageLink(4, mainDataAndimages_urls[index_main][4]))
                        setCellValue(sheet.cell(row, 22), getProperImageLink(5, mainDataAndimages_urls[index_main][4]))
                        setCellValue(sheet.cell(row, 23), str_video_url)
                        setCellValue(sheet.cell(row, 24), base_row_value[3])
                        row = row + 1
                        index_main = index_main + 1
                    
                    self.countChanged.emit(round(count))
                    count += 100/length
                self.pageChanged.emit(page_number+1)
            workbook.save(excel_file)
            self.countFinished.emit()
        except:
            self.countErrored.emit()
        
class Ui_MainWindow(object):
    def setupUi(self, MainWindow):   

    
        ## MainWind Handle
        MainWindow.setObjectName("MainWindow")
        MainWindow.setEnabled(True)
        MainWindow.setFixedSize(850, 360)
        MainWindow.setBaseSize(QtCore.QSize(600, 400))
        ##
        
        ## Utils
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(120, 30, 641, 31))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(18)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setLayoutDirection(QtCore.Qt.RightToLeft)
        self.label.setObjectName("label")
        ##
        
        ## ProgreeBar Handle
        self.progressBar = QtWidgets.QProgressBar(self.centralwidget)
        self.progressBar.setGeometry(QtCore.QRect(30, 240, 815, 21))
        self.progressBar.setProperty("value", 0)
        self.progressBar.setTextDirection(QtWidgets.QProgressBar.TopToBottom)
        self.progressBar.setObjectName("progressBar")
        ##
        
        
        ## PushButton Handle
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(330, 280, 211, 51))
        font = QtGui.QFont()
        font.setFamily("Sitka Heading")
        font.setPointSize(14)
        font.setBold(False)
        font.setWeight(50)
        self.pushButton.setFont(font)
        self.pushButton.setObjectName("pushButton")
        #self.pushButton.clicked.connect(self.doAction)
        self.pushButton.clicked.connect(self.onButtonClick)
        ##
        
        
        ## List Handle
        self.listView = QtWidgets.QListView(self.centralwidget)
        self.listView.setGeometry(QtCore.QRect(30, 130, 781, 101))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.listView.setFont(font)
        self.listView.setObjectName("listView")
        self.listView.setModel(model)
        item = QtGui.QStandardItem(entries[0])
        model.appendRow(item)
        self.listView.setEnabled(False)
        ##
        
        
        ## product check box handle
        font = QtGui.QFont()
        font.setPointSize(12)
        self.chk_proudcts = QtWidgets.QCheckBox(self.centralwidget)
        self.chk_proudcts.setGeometry(QtCore.QRect(30, 90, 111, 31))
        self.chk_proudcts.setFont(font)
        self.chk_proudcts.setChecked(True)
        self.chk_proudcts.setObjectName("chk_proudcts")
        self.chk_proudcts.stateChanged.connect(self.on_state_changed) 
        
        ## parts check box handle
        self.chk_parts = QtWidgets.QCheckBox(self.centralwidget)
        self.chk_parts.setEnabled(False)
        self.chk_parts.setGeometry(QtCore.QRect(190, 90, 151, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.chk_parts.setFont(font)
        self.chk_parts.setObjectName("chk_parts")
        
        ## retailer check box handle
        self.chk_retailer = QtWidgets.QCheckBox(self.centralwidget)
        self.chk_retailer.setEnabled(False)
        self.chk_retailer.setGeometry(QtCore.QRect(400, 90, 91, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.chk_retailer.setFont(font)
        self.chk_retailer.setObjectName("chk_retailer")
        
        ## about check box handle
        self.chk_about = QtWidgets.QCheckBox(self.centralwidget)
        self.chk_about.setEnabled(False)
        self.chk_about.setGeometry(QtCore.QRect(550, 90, 141, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.chk_about.setFont(font)
        self.chk_about.setObjectName("chk_about")
        
        ## chk_contact check box handle
        self.chk_contact = QtWidgets.QCheckBox(self.centralwidget)
        self.chk_contact.setEnabled(False)
        self.chk_contact.setGeometry(QtCore.QRect(740, 90, 91, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.chk_contact.setFont(font)
        self.chk_contact.setObjectName("chk_contact")
        ##
        
        ##
        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.actionOpen = QtWidgets.QAction(MainWindow)
        self.actionOpen.setObjectName("actionOpen")
        self.actionExit = QtWidgets.QAction(MainWindow)
        self.actionExit.setObjectName("actionExit")

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
    
    def onButtonClick(self):
        if(model.rowCount()==0):
            QMessageBox.about(MainWindow, "ERROR", "No URL selected!")
            return
        self.pushButton.setText('Extracting...(1/4)')
        self.progressBar.setValue(0)
        self.calc = External()
        self.calc.countChanged.connect(self.onCountChanged)
        self.calc.countFinished.connect(self.onThreadFinished)
        self.calc.countErrored.connect(self.onCountErrored)
        self.calc.pageChanged.connect(self.onPageChanged)
        self.calc.start()

    def onPageChanged(self, value):
        self.pushButton.setText('Extracting...(' + str(value) + '/4)' )
        self.progressBar.setValue(0)

    def onCountChanged(self, value):
        self.progressBar.setValue(value)
        
    def onThreadFinished(self):
        self.progressBar.setValue(100)
        self.pushButton.setText('Start')
        QMessageBox.about(MainWindow, "SUCCESS", "Scraping Completed!")

    def onCountErrored(self):
        QMessageBox.about(MainWindow, "ERROR", "Check the Internet Connection!\n Or Check whether Excel is opened.")
        return
        
    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "BradWebScraper"))
        self.label.setText(_translate("MainWindow", "Shop Our Selection of Vacuum Products at MetroVac Today!"))
        self.pushButton.setText(_translate("MainWindow", "START"))
        self.chk_proudcts.setText(_translate("MainWindow", "PRODUCTS"))
        self.chk_parts.setText(_translate("MainWindow", "Parts/Accessories"))
        self.chk_retailer.setText(_translate("MainWindow", "Retailers"))
        self.chk_about.setText(_translate("MainWindow", "About Metrovac"))
        self.chk_contact.setText(_translate("MainWindow", "Contact"))
        self.actionOpen.setText(_translate("MainWindow", "Open"))
        self.actionExit.setText(_translate("MainWindow", "Exit"))

    def on_state_changed(self, state):
        model = self.listView.model()
        if state == QtCore.Qt.Checked:
            item = QtGui.QStandardItem(entries[0])
            model.appendRow(item)
        else:
            model.removeRow(0)

if __name__ == "__main__":
    import sys
    import time
    import requests
    from bs4 import BeautifulSoup
    import urllib
    import urllib.request
    from urllib.request import Request, urlopen
    import ssl
    import csv
    import os
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    import json
    
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
